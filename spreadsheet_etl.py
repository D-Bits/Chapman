"""
A basic Python-based ETL program for working with spreadsheet data sources
"""
from sqlalchemy import create_engine
import psycopg2, os, openpyxl as xl
from openpyxl.chart import Reference


# Create custom connection function
def db_connection():

    # Define credentials using environment vars
    host = os.environ.get('POSTGRES_HOST')
    db = 'ETL_Test'
    user = os.environ.get('POSTGRES_USER')
    pgpass = os.environ.get('PG_PASS')

    connection_params = psycopg2.connect(database=db, host=host, user=user, password=pgpass)
    return connection_params


# Define engine
engine = create_engine('postgresql+psycopg2://', creator=db_connection)


def get_data():

    # Specify workbork & sheet
    wb = xl.load_workbook('world-ports.xlsx')
    sheet = wb['ports']
    


# Define ETL operations
def etl():

    # Specify workbork & sheet
    wb = xl.load_workbook('world-ports.xlsx')
    sheet = wb['ports']
    # specific rows (and columns) to be inserted.
    sheet_range = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=1, max_col=4)

    for row in range(1, sheet.max_row + 1):
        engine.execute(
            "INSERT INTO ports(city, nation, MetroPop, AnnualCargo) VALUES(%s, %s, %s, %s)",
            row
        )


def main():

    etl()


main()
